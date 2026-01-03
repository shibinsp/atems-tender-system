"""
Export Service for ATEMS
Handles exporting data to PDF, Excel, and CSV formats
"""

import io
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

# Try to import export libraries
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to various formats"""

    def __init__(self):
        self.styles = None
        if PDF_AVAILABLE:
            self.styles = getSampleStyleSheet()
            self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom PDF styles"""
        if not PDF_AVAILABLE:
            return

        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1e3a5f')
        ))

        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=12,
            spaceAfter=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666')
        ))

        # Section header style
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#1e3a5f')
        ))

    def export_to_csv(self, data: List[Dict], filename: str = "export.csv") -> io.BytesIO:
        """
        Export data to CSV format

        Args:
            data: List of dictionaries containing the data
            filename: Name for the CSV file

        Returns:
            BytesIO buffer containing the CSV data
        """
        if not data:
            raise ValueError("No data to export")

        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        # Convert to bytes
        output = io.BytesIO()
        output.write(buffer.getvalue().encode('utf-8'))
        output.seek(0)
        return output

    def export_to_excel(
        self,
        data: List[Dict],
        sheet_name: str = "Data",
        title: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export data to Excel format

        Args:
            data: List of dictionaries containing the data
            sheet_name: Name for the Excel sheet
            title: Optional title for the report

        Returns:
            BytesIO buffer containing the Excel file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Excel export unavailable.")

        if not data:
            raise ValueError("No data to export")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        start_row = 1

        # Add title if provided
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data[0].keys()))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(bold=True, size=16, color="1e3a5f")
            title_cell.alignment = Alignment(horizontal="center")
            start_row = 3

        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(start_row, len(data) + start_row + 1):
                try:
                    cell_value = str(ws.cell(row=row, column=col).value or '')
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_tender_summary_pdf(
        self,
        tender_data: Dict,
        bids: List[Dict],
        evaluation_type: str
    ) -> io.BytesIO:
        """
        Export tender summary as PDF

        Args:
            tender_data: Tender information
            bids: List of bid data
            evaluation_type: Type of evaluation (L1, T1, QCBS)

        Returns:
            BytesIO buffer containing the PDF
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not installed. PDF export unavailable.")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        elements = []

        # Header
        elements.append(Paragraph("AI-Based Tender Evaluation & Management System", self.styles['CustomTitle']))
        elements.append(Paragraph("TENDER SUMMARY REPORT", self.styles['CustomSubtitle']))
        elements.append(Spacer(1, 20))

        # Tender Details Section
        elements.append(Paragraph("Tender Information", self.styles['SectionHeader']))

        tender_info = [
            ["Tender ID:", tender_data.get('tender_id', 'N/A')],
            ["Title:", tender_data.get('title', 'N/A')],
            ["Category:", tender_data.get('category', 'N/A')],
            ["Estimated Value:", f"₹ {tender_data.get('estimated_value', 0):,.2f}"],
            ["Evaluation Type:", evaluation_type],
            ["Submission Deadline:", tender_data.get('submission_deadline', 'N/A')],
            ["Status:", tender_data.get('status', 'N/A')]
        ]

        tender_table = Table(tender_info, colWidths=[2*inch, 4*inch])
        tender_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(tender_table)
        elements.append(Spacer(1, 20))

        # Bids Summary Section
        if bids:
            elements.append(Paragraph("Bids Summary", self.styles['SectionHeader']))

            bid_headers = ["Rank", "Bidder", "Financial Amount", "Technical Score", "Status"]
            bid_data = [bid_headers]

            for bid in bids:
                bid_data.append([
                    str(bid.get('rank', '-')),
                    bid.get('bidder_name', 'N/A'),
                    f"₹ {bid.get('financial_amount', 0):,.2f}",
                    f"{bid.get('technical_score', 0):.2f}",
                    bid.get('status', 'N/A')
                ])

            bid_table = Table(bid_data, colWidths=[0.7*inch, 2*inch, 1.5*inch, 1.2*inch, 1.2*inch])
            bid_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')])
            ]))
            elements.append(bid_table)

        elements.append(Spacer(1, 30))

        # Footer
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ATEMS - Government of India"
        elements.append(Paragraph(footer_text, self.styles['Normal']))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def export_comparative_statement_pdf(
        self,
        tender_data: Dict,
        bids: List[Dict],
        evaluation_criteria: List[Dict],
        recommendation: str
    ) -> io.BytesIO:
        """
        Export comparative statement as PDF

        Args:
            tender_data: Tender information
            bids: List of bid data with scores
            evaluation_criteria: List of evaluation criteria
            recommendation: Evaluation recommendation

        Returns:
            BytesIO buffer containing the PDF
        """
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not installed. PDF export unavailable.")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=30,
            bottomMargin=30
        )

        elements = []

        # Header
        elements.append(Paragraph("COMPARATIVE STATEMENT OF BIDS", self.styles['CustomTitle']))
        elements.append(Spacer(1, 10))

        # Tender Info
        tender_info_text = f"""
        <b>Tender:</b> {tender_data.get('title', 'N/A')}<br/>
        <b>Tender ID:</b> {tender_data.get('tender_id', 'N/A')}<br/>
        <b>Estimated Value:</b> ₹ {tender_data.get('estimated_value', 0):,.2f}
        """
        elements.append(Paragraph(tender_info_text, self.styles['Normal']))
        elements.append(Spacer(1, 20))

        # Comparative Table
        if bids:
            # Build headers dynamically
            headers = ["S.No", "Bidder Name", "Financial Quote"]
            for criteria in evaluation_criteria[:3]:  # Limit to 3 criteria for space
                headers.append(criteria.get('name', 'Criteria')[:15])
            headers.extend(["Total Score", "Rank"])

            table_data = [headers]

            for idx, bid in enumerate(sorted(bids, key=lambda x: x.get('rank', 999)), 1):
                row = [
                    str(idx),
                    bid.get('bidder_name', 'N/A')[:20],
                    f"₹{bid.get('financial_amount', 0):,.0f}"
                ]
                # Add criteria scores
                scores = bid.get('criteria_scores', {})
                for criteria in evaluation_criteria[:3]:
                    score = scores.get(criteria.get('id'), '-')
                    row.append(str(score) if score != '-' else '-')
                row.extend([
                    f"{bid.get('combined_score', 0):.1f}",
                    str(bid.get('rank', '-'))
                ])
                table_data.append(row)

            # Calculate column widths
            num_cols = len(headers)
            col_width = 7.5 * inch / num_cols

            comp_table = Table(table_data, colWidths=[col_width] * num_cols)
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')])
            ]))
            elements.append(comp_table)

        elements.append(Spacer(1, 20))

        # Recommendation
        elements.append(Paragraph("Recommendation", self.styles['SectionHeader']))
        elements.append(Paragraph(recommendation, self.styles['Normal']))

        elements.append(Spacer(1, 30))

        # Signature Section
        sig_text = """
        <br/><br/>
        _________________________ &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; _________________________<br/>
        <b>Evaluator Signature</b> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <b>Chairperson Signature</b>
        """
        elements.append(Paragraph(sig_text, self.styles['Normal']))

        # Footer
        footer = f"<br/><br/>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ATEMS"
        elements.append(Paragraph(footer, self.styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def export_audit_trail_excel(
        self,
        audit_logs: List[Dict],
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> io.BytesIO:
        """
        Export audit trail to Excel

        Args:
            audit_logs: List of audit log entries
            start_date: Optional start date filter
            end_date: Optional end date filter

        Returns:
            BytesIO buffer containing the Excel file
        """
        title = f"Audit Trail Report"
        if start_date and end_date:
            title += f" ({start_date} to {end_date})"

        return self.export_to_excel(audit_logs, "Audit Trail", title)

    def export_bids_matrix_excel(
        self,
        tender_title: str,
        bids: List[Dict],
        criteria: List[Dict]
    ) -> io.BytesIO:
        """
        Export bid evaluation matrix to Excel

        Args:
            tender_title: Title of the tender
            bids: List of bids with evaluation data
            criteria: Evaluation criteria

        Returns:
            BytesIO buffer containing the Excel file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Excel export unavailable.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluation Matrix"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        subheader_fill = PatternFill(start_color="4a5568", end_color="4a5568", fill_type="solid")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Bid Evaluation Matrix - {tender_title}"
        ws['A1'].font = Font(bold=True, size=14, color="1e3a5f")

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)

        # Headers row
        row = 4
        headers = ["Bidder", "Financial Amount"]
        for c in criteria:
            headers.append(c.get('name', 'Criteria'))
        headers.extend(["Technical Score", "Financial Score", "Combined Score", "Rank"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        for bid in sorted(bids, key=lambda x: x.get('rank', 999)):
            row += 1
            ws.cell(row=row, column=1, value=bid.get('bidder_name', 'N/A'))
            ws.cell(row=row, column=2, value=bid.get('financial_amount', 0))

            col = 3
            scores = bid.get('criteria_scores', {})
            for c in criteria:
                ws.cell(row=row, column=col, value=scores.get(c.get('id'), 0))
                col += 1

            ws.cell(row=row, column=col, value=bid.get('technical_score', 0))
            ws.cell(row=row, column=col + 1, value=bid.get('financial_score', 0))
            ws.cell(row=row, column=col + 2, value=bid.get('combined_score', 0))
            ws.cell(row=row, column=col + 3, value=bid.get('rank', '-'))

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# Singleton instance
export_service = ExportService()
